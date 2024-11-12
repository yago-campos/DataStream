import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Passo 1: Ler a planilha original
file_path = 'PENDENTES.xlsx'
df = pd.read_excel(file_path)

# Passo 2: Filtrar a coluna "Retorno do distribuidor"

status_aceitos = [
    'ACEITO COM SUCESSO', 
    'AGUARDANDO RETORNO DO DISTRIBUIDOR',
    'PRODUTO ACEITO COM SUCESSO',
    'PRODUTO PARCIALMENTE ACEITO'
]

filtered_df = df[df['Retorno do distribuidor'].isin(status_aceitos)]

# Passo 3: Remover linhas da coluna "Nota recebida" com valor "1"
filtered_df = filtered_df[filtered_df['Nota recebida'] != 1]

# Passo 4: Criar planilhas separadas para cada distribuidor e tipo (CA/WE)
distribuidores = filtered_df['Nome do distribuidor'].unique()

# Função para adicionar a aba "RESUMO" com a tabela dinâmica
def add_pivot_sheet(writer, df, sheet_name):
    # Adicionar os dados à planilha
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Criar tabela dinâmica usando pandas
    pivot_df = pd.pivot_table(
        df,
        values='Pedido líquido',  # Substitua pelo nome da coluna que representa os valores
        index=['Nome do distribuidor', 'Retorno do distribuidor'],
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    # Adicionar a aba "RESUMO"
    resumo_sheet = workbook.create_sheet('RESUMO')

    # Escrever a tabela dinâmica na aba "RESUMO"
    for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True)):
        for c_idx, value in enumerate(row):
            resumo_sheet.cell(row=r_idx+1, column=c_idx+1, value=value)
    
    print(f'Tabela dinâmica adicionada na aba RESUMO para {sheet_name}')

# Salvar planilhas separadas
for distribuidor in distribuidores:
    # Filtrar por distribuidor
    distribuidor_df = filtered_df[filtered_df['Nome do distribuidor'] == distribuidor]

    # Filtrar por tipo (CA ou WE)
    ca_df = distribuidor_df[distribuidor_df['Origem do pedido'] == 'Canal Autorizador']
    we_df = distribuidor_df[distribuidor_df['Origem do pedido'] != 'Canal Autorizador']

    if not ca_df.empty:
        ca_file_name = f'PENDENCIAS ({distribuidor}) - CA.xlsx'
        with pd.ExcelWriter(ca_file_name, engine='openpyxl') as writer:
            add_pivot_sheet(writer, ca_df, 'Dados')
        print(f'Arquivo salvo: {ca_file_name}')
    
    if not we_df.empty:
        we_file_name = f'PENDENCIAS ({distribuidor}) - WE.xlsx'
        with pd.ExcelWriter(we_file_name, engine='openpyxl') as writer:
            add_pivot_sheet(writer, we_df, 'Dados')
        print(f'Arquivo salvo: {we_file_name}')
